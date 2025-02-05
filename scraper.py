from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

from openpyxl import Workbook, load_workbook


class Scraper:
    def __init__(self,
                 start_page_num: int,
                 min_wait_time: int,
                 max_wait_time: int,
                 id2country: dict[str, str],
                 headers: list[str],
                 xlsx_path: str,
                 xlsx_page_title: str) -> None:

        self.website = "https://portal.eaeunion.org/sites/commonprocesses/ru-ru/Pages/DrugRegistrationDetails.aspx"

        # Count IDs
        self.count_db_id = "ComboBox1-input"
        self.count_option_id = "ComboBox1-list4"

        # Pages classes
        self.current_page_class = "ecc-page-number-input"
        self.last_page_class = "eec-page-count"
        self.next_bt_class = "arrow-right"

        # Page
        self.start_page_num = start_page_num
        self.current_page_num = 1
        self.__last_page_num = None

        # Wait time
        self.min_wait_time = min_wait_time
        self.max_wait_time = max_wait_time

        # Countries codes
        self.id2country = id2country

        # xlsx
        self.headers = headers
        self.xlsx_path = xlsx_path
        self.xlsx_page_title = xlsx_page_title

    def create_xlsx(self) -> None:
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = self.xlsx_page_title

        # Write headers
        ws.append(self.headers)

        # Save the workbook
        wb.save(self.xlsx_path)

    def write_to_xlsx(self,
                      data: dict[str, list[str]]) -> None:
        wb = load_workbook(self.xlsx_path)
        ws = wb.active

        # Write new entries
        rows = zip(*data.values())
        for row in rows:
            ws.append(row)

        # Save the workbook
        wb.save(self.xlsx_path)

    def load_driver(self) -> None:
        self.driver = webdriver.Edge()
        self.driver.implicitly_wait(self.max_wait_time)
        self.driver.get(self.website)

    @property
    def last_page_num(self) -> int:
        if self.__last_page_num is None:
            last_page_present = EC.presence_of_element_located(
                (By.CLASS_NAME, self.last_page_class))
            last_page = WebDriverWait(
                self.driver, self.max_wait_time).until(last_page_present)

            self.__last_page_num = int(last_page.text)

        return self.__last_page_num

    def change_num_entries(self) -> None:
        # Dropdown
        count_db_present = EC.presence_of_element_located(
            (By.ID, self.count_db_id))
        count_db = WebDriverWait(
            self.driver, self.max_wait_time).until(count_db_present)
        # count_db = self.driver.find_element(By.ID, self.count_db_id)
        count_db.click()

        # Option
        count_option_present = EC.element_to_be_clickable(
            (By.ID, self.count_option_id))
        count_option = WebDriverWait(
            self.driver, self.min_wait_time).until(count_option_present)
        count_option.click()

    def wait_page_loading(self) -> None:
        page_num_present = EC.text_to_be_present_in_element_attribute((By.CLASS_NAME, "ecc-page-number-input"),
                                                                      "placeholder",
                                                                      str(self.current_page_num))
        WebDriverWait(self.driver, self.max_wait_time).until(
            page_num_present)

    def change_page_num(self) -> None:
        page_num_present = EC.presence_of_element_located(
            (By.CLASS_NAME, self.current_page_class))
        page_num = WebDriverWait(
            self.driver, self.max_wait_time).until(page_num_present)
        page_num.send_keys(str(self.start_page_num))
        page_num.send_keys(Keys.RETURN)

        self.current_page_num = self.start_page_num

        # Wait until new page is loaded
        self.wait_page_loading()

    def paginate(self) -> None:
        self.driver.implicitly_wait(self.max_wait_time)

        if self.current_page_num < self.last_page_num:
            next_bt_present = EC.presence_of_element_located(
                (By.CLASS_NAME, self.next_bt_class))
            next_bt = WebDriverWait(
                self.driver, self.max_wait_time).until(next_bt_present)
            next_bt.click()

            self.current_page_num += 1

            # Wait until new page is loaded
            self.wait_page_loading()

        else:
            self.current_page_num += 1

    def get_cell_text(self, cell) -> str:
        self.driver.implicitly_wait(0)
        div = cell.find_element(By.XPATH, ".//div")
        # country_spans = div.find_elements(
        #     By.XPATH, ".//spa n[contains(@class, 'i-country--')]")

        country_spans = div.find_elements(By.TAG_NAME, "span")

        if country_spans:
            results = []
            for span in country_spans:
                # Extract country code from class
                country_code = span.get_attribute(
                    "class").split("--")[-1].split()[0]
                country_name = self.id2country.get(country_code, country_code)

                # Get immediately following text using JavaScript
                text = cell.parent.execute_script(
                    "return arguments[0].nextSibling.textContent.trim();", span)

                results.append(f"{country_name} - {text}")
            return ", ".join(results)

        # Simple case with no country codes
        return div.text.strip()

    def get_data(self) -> dict:
        # Init dict with empty lists
        data = {k: [] for k in self.headers}

        # Get table and rows
        table = self.driver.find_element(By.XPATH, "//tbody")
        rows = table.find_elements(By.TAG_NAME, "tr")

        # Iterate through rows
        for row in rows:
            cells = row.find_elements(By.TAG_NAME, "td")
            values = [cell.text for cell in cells]
            values[0] = self.get_cell_text(cells[0])
            for key, value in zip(data.keys(), values):
                data[key].append(value)

        return data

    def scrape(self) -> None:
        # Load driver
        self.load_driver()

        # Change number of entries per page
        self.change_num_entries()

        # Change current page num
        if self.start_page_num != self.current_page_num:
            self.change_page_num()

        # Create empty xlsx
        if self.start_page_num == 1:
            self.create_xlsx()

        while self.current_page_num <= self.last_page_num:
            print(f"Scraping {self.current_page_num} page...")
            data = self.get_data()
            self.write_to_xlsx(data)

            self.paginate()

        print("Scraping has finished.")
